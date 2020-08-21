#импортируем необходимые библиотеки
from colorama import init #для стиля ветового
from openpyxl.styles import Font, Color, colors
from openpyxl import load_workbook
from tkinter import *  #import library
from tkinter import messagebox
init()

def write(a,b,c):
    #для удобства вводим название файла
    filename = "Список.xlsx"
    filename_1 = "Шаблон_Сегмента.xlsx"

    #вытянули данные с документа
    active_excel = load_workbook(filename=filename,data_only=True)#data_only=True
    active_excel_1 = load_workbook(filename=filename_1,data_only=True)#data_only=True

    #делаем или смотрим активный лист
    active_sheet = active_excel.active
    active_sheet_1 = active_excel_1.active

    #нужно понять максимальный размер данных на листе
    max_row = active_sheet.max_row
    max_row_1 = active_sheet_1.max_row

    #корректировка регистра ввода
    a = a.upper()
    c = c.lower()

    #поиск
    #делаем цикл по заполнению СЦЕПИТЬ всего диапазона колонки в нашем списке
    spisok = []
    for i in range(2,(max_row + 1)):
        a_s = active_sheet["A" + str(i)].value
        b_s = active_sheet["B" + str(i)].value
        g_s = active_sheet["G" + str(i)].value

        d_s = str(a_s) + str(g_s) + str(b_s)
        spisok.append(d_s)

    spisok_1 = [] #Номенклатура A (в списке)
    spisok_2 = [] #Характеристика B (в списке)
    spisok_2_1 = [] #пробелы
    spisok_2_2 = [] #пробелы
    spisok_3 = [] #Единица хранения C (в списке)
    spisok_4 = [] #ТипИзделия D (в списке)
    spisok_5 = [] #Штрихкод E (в списке)
    spisok_all = [spisok_1,spisok_2,spisok_2_1,spisok_2_2,spisok_3,spisok_4,spisok_5]

    display = []
    e = 0
    for q in spisok:
        f = spisok[e][:]
        if a in f:
            if b in f:
                if c in f:
                    display.append(str(e+2) + ' ' + f + "\n")
                    print(str(e+2) + ' ' + f)
                    A = active_sheet["A" + str(e+2)].value
                    spisok_1.append(A)
                    B = active_sheet["B" + str(e+2)].value
                    spisok_2.append(B)
                    C = active_sheet["C" + str(e+2)].value
                    spisok_3.append(C)
                    D = active_sheet["D" + str(e+2)].value
                    spisok_4.append(D)
                    E = active_sheet["E" + str(e+2)].value
                    spisok_5.append(E)
                    spisok_2_1.append(" ")
                    spisok_2_2.append(" ")
        e += 1

    #запишем полученные данные в файл заливки сегмента
    for g in range (len(spisok_1)):
        for h in range (len(spisok_all)):
            _=active_sheet_1.cell(column=h+1, row=max_row_1+g+1, value=spisok_all[h][g])

    # далее отформатируем таблицу
    style_1 = Font(name='TimesNewRoman', color=colors.BLACK,
               bold=False, size=14)#underline='double'
    max_row_2=active_sheet_1.max_row
    for z in range(2, (max_row_2 + 1)):
        a = active_sheet_1['A' + str(z)]
        b = active_sheet_1['B' + str(z)]
        e = active_sheet_1['E' + str(z)]
        f = active_sheet_1['F' + str(z)]
        g = active_sheet_1['G' + str(z)]

        a.font = style_1
        b.font = style_1
        e.font = style_1
        f.font = style_1
        g.font = style_1

    # сохраняем изменения
    active_excel_1.save("Шаблон_Сегмента.xlsx") #сохраняем все изменения

    #выводим сообщение о успешности операции
    # выведем записанный товар во всплывающем окне
    messagebox.showinfo("ДАННЫЕ ЗАПИСАНЫ И СОХРАНЕНЫ!", display)

def MyForm():
    """Функция показывает форму"""
    # делаем формы ввода (сделать в виде форм)
    root = Tk()
    root.title("ФОРМИРОВАНИЕ СЕГМЕНТА")

    #определяем переменные
    group = StringVar()
    name = StringVar()
    characteristic = StringVar()

    # определили полей Label и Entry
    mylabel1 = Label(root, text="ВВЕДИТЕ ГРУППУ:", padx=30)
    a1 = Entry(root, width=25, borderwidth=3, textvariable=group)

    mylabel2 = Label(root, text="ВВЕДИТЕ ЧАСТЬ ИМЕНИ:", padx=30)
    b1 = Entry(root, width=25, borderwidth=3, textvariable=name)

    mylabel3 = Label(root, text="ВВЕДИТЕ ХАРАКТЕРИСТИКУ:", padx=30)
    c1 = Entry(root, width=25, borderwidth=3, textvariable=characteristic)

    # расположение label
    mylabel1.grid(row=0, column=0)
    mylabel2.grid(row=1, column=0)
    mylabel3.grid(row=2, column=0)

    # расположение Entry
    a1.grid(row=0, column=1)
    b1.grid(row=1, column=1)
    c1.grid(row=2, column=1)

    def get_text():
        """Получает введенный текст и передает его"""
        a = group.get()
        b = name.get()
        c = characteristic.get()
        # удаляем предыдущие данные
        c1.delete(0, END)
        b1.delete(0, END)
        a1.delete(0, END)
        #передаем данные для обработки
        write(a, b, c)

    #кнопки Записать, Выйти и их расположение
    MyButton = Button(root, text="ЗАПИСАТЬ СЕГМЕНТ", padx=30, fg="BLUE", bg="YELLOW", borderwidth=3, command=lambda: get_text())
    MyButton.grid(row=4, column=1)

    MyButton = Button(root, text="ВЫЙТИ", padx=70, fg="YELLOW", bg="BLUE", borderwidth=3, command=lambda: quit())
    MyButton.grid(row=4, column=0)

    root.mainloop()

while 1 == 1:
    MyForm()

input()